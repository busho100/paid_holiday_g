import streamlit as st
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime

# Excel文書を作成する関数
def create_excel(data):
    wb = openpyxl.load_workbook('template.xlsx')
    ws = wb.active

    # データをExcelシートに追加
    ws["H4"] = data["申請日"]
    ws["B8"] = data["所属"]
    ws["B10"] = data["氏名"]
    ws["B13"] = data["開始日"]
    ws["B14"] = data["終了日"]
    ws["G13"] = data["日数"]
    ws["G13"].alignment = Alignment(horizontal='center', vertical='center')
    ws["R13"].alignment = Alignment(horizontal='center', vertical='center')
    ws["B16"] = data["区別"]
    ws["B16"].alignment = Alignment(horizontal='center', vertical='center')
    ws["M16"].alignment = Alignment(horizontal='center', vertical='center')
    ws["B17"] = data["事由"]
    ws["B17"].alignment = Alignment(horizontal='center', vertical='center')
    ws["M17"].alignment = Alignment(horizontal='center', vertical='center')
    ws["B18"] = data["備考"]
    ws["B18"].alignment = Alignment(horizontal='center', vertical='center')
    ws["M18"].alignment = Alignment(horizontal='center', vertical='center')

    # ファイル名を生成
    filename = f"休暇届_{data['申請日']}_{data['所属']}_{data['氏名']}.xlsx"
    wb.save(filename)
    return filename

# Streamlit UI
st.title("業務改善アプリ")
st.header("有給申請")

# ユーザー入力を収集
name = st.text_input("氏名")
request_date = st.date_input("申請日")
formatted_request_date = f"令和{request_date.year - 2018}年{request_date.month}月{request_date.day}日"
department = st.selectbox("所属", ["総務", "経理", "開発", "広報", "給食", "営繕", "営業"])
start_date = st.date_input("開始日")
formatted_start_date = f"令和{start_date.year - 2018}年{start_date.month}月{start_date.day}日"
end_date = st.date_input("終了日")
formatted_end_date = f"令和{end_date.year - 2018}年{end_date.month}月{end_date.day}日"
days = st.number_input("日数", min_value=0.5, format="%.1f")
leave_type = st.selectbox("区別", ["①有給休暇", "②生理休暇", "③慶弔休暇", "④産前産後休暇", "⑤特別休暇", "⑥その他"])
reason = st.text_area("事由")
remarks = st.text_area("備考")

# 申請ボタン
if st.button("申請書作成"):
    data = {
        "申請日": formatted_request_date,
        "所属": department,
        "氏名": name,
        "開始日": formatted_start_date,
        "終了日": formatted_end_date,
        "日数": days,
        "区別": leave_type,
        "事由": reason,
        "備考": remarks
    }
    filename = create_excel(data)
    st.success("申請書が作成されました。")

    # ダウンロードリンクの提供
    with open(filename, "rb") as file:
        st.download_button("申請書をダウンロード", data=file, file_name=filename)

